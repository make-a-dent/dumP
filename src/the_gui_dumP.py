# -*- coding: utf-8 -*-
# @Time    : 2019/5/24 15:07
# @Author  : zhouqiang
# @Site    : 
# @File    : the_gui_dumP.py
# @Software: PyCharm

import pymysql
import openpyxl
from openpyxl.styles import Font, colors, Alignment, Border, Side

import tkinter
import tkinter.messagebox as t_m
import tkinter.font as tk_font
import re
import os
from functools import wraps

BASE_DIR = ''

CATALOG_COLS = {'库': 1, '表': 2, 'COMMENT': 3}

MY_COLS = ['列名', '类型', '默认值', 'COMMENT', '字符集', '排序规则']
COLS_POS = {
    '列名': 0, '类型': 1, '默认值': 2, 'COMMENT': 3, '字符集': 4, '排序规则': 5
}

BOLD_RED_FONT = Font(color=colors.RED, size=12, bold=True)
COMMON_CELL_FONT = Font(size=12)
FONT_MM = 1.3  # 12号，小四，4.23mm
CENTER_H_V = Alignment(horizontal='center', vertical='center')


def sys_exit(err_msg):
    print(err_msg)
    # sys.exit(1)
    raise RuntimeError(err_msg)


def get_mysql_conn(host, port, user, password, database):
    print('host: %s, port: %s, user: %s, password: %s, database: %s' % (host, port, user, password, database))
    return pymysql.connect(host=host, port=port, user=user, password=password, database=database)


def set_conn_v2(func):
    """一个简单的，设置conn装饰器，如果字典参数中没有conn，就自动加一个进去，先不管db是谁了，用默认的吧"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        conn = kwargs.get('conn')
        close_it = False
        if conn is None:
            conn = get_mysql_conn(host=kwargs.get('host'), port=kwargs.get('port'), user=kwargs.get('user'),
                                  password=kwargs.get('password'), database=kwargs.get('db_str'))
            close_it = True
        kwargs.update({'conn': conn})
        try:
            result = func(*args, **kwargs)
        finally:
            if close_it:
                conn.close()
        return result
    return wrapper


@set_conn_v2
def read(db_str, table_str, sql_list=None, **kwargs):
    print('库：%s，表：%s。' % (db_str, table_str))
    conn = kwargs.get('conn')
    close_this_conn = kwargs.get('close_this_conn', False)

    sql_str = None
    if isinstance(sql_list, str):
        sql_str = sql_list
    elif sql_list is None:
        sql_str = 'select * from %s' % table_str
    elif isinstance(sql_list, list):
        sql_str = ' '.join(sql_list)
    print('将要被送往服务器的SQL：%s' % sql_str)
    cursor = conn.cursor()
    cursor.execute(sql_str)
    all_data = cursor.fetchall()
    # print(all_data)
    print('是不是需要关闭连接：%s' % close_this_conn)
    if close_this_conn:
        conn.close()
    return all_data


def process_special(pattern, x):
    res = re.search(pattern, x)
    character = ''
    if res:
        print('特殊处理的字符串：🚁%s🚁' % x)
        c_set = res.group()
        character = c_set.split(' ')[-1]
        print('得到的值：%s' % character)
    return character


def deal_a_row(x):
    print('处理前：🍼%s🍼' % x)
    x_res = ['' for _ in range(len(MY_COLS))]

    # 处理 单引号内容中含有空格的情况
    pattern = r'\'(\S+)(( (\S+))+)\''
    res = re.search(pattern, x)
    if res:
        print('单引号中内容处理，之前：🍭%s🍭' % x)
        c_set = res.group()
        x = x.replace(c_set, ','.join(c_set.split(' ')))
        print('单引号中内容处理，之后：🍭%s🍭' % x)

    x = x.replace('\"', '')
    x = x.replace('\'', '')
    x = x.replace('`', '')
    x = x.replace(' ' * 2, '')

    # 先做这些，就能判断这一行描述的是什么
    x = x.replace('DEFAULT NULL', 'DEFAULT-NULL')
    x = x.replace('NOT NULL AUTO_INCREMENT', 'NOT-NULL-AUTO_INCREMENT')
    x = x.replace('NOT NULL', 'NOT-NULL')
    x = x.replace('CHARACTER SET', 'CHARACTER-SET')
    x = x.replace('PRIMARY KEY', '主键')
    x = x.replace('KEY', '索引')  # 这个一定要在上面的主键替换之后
    x = x.replace(') ENGINE', 'ENGINE')
    x = x.replace('USING BTREE', 'BTREE')
    x = x.replace('DEFAULT CHARSET', 'DEFAULT-CHARSET')

    if x.startswith('主键') or x.startswith('索引') or x.startswith('ENGINE'):
        x_res = x.split()
    else:
        # 处理 列名
        d_type = '列名'
        d_value = x.split()[0]
        x_res[COLS_POS[d_type]] = d_value

        # 处理 类型
        d_type = '类型'
        d_value = x.split()[1]
        x_res[COLS_POS[d_type]] = d_value

        # 处理 默认值
        d_type = '默认值'
        d_value = ''
        if re.search(r'DEFAULT-NULL', x):
            d_value = '空'
        elif re.search(r'NOT-NULL-AUTO_INCREMENT', x):
            d_value = '非空+自增'
        elif re.search(r'NOT-NULL', x):
            d_value = '非空'
        x_res[COLS_POS[d_type]] = d_value

        # 处理 COMMENT
        d_type = 'COMMENT'
        d_value = ''
        res = re.search(r'COMMENT (\S+)', x)
        # res = re.search(r'COMMENT (\w+)', x)
        if res:
            res = res.group()
            print('找到一个COMMENT：%s' % res)
            d_value = res.split()[1]
        x_res[COLS_POS[d_type]] = d_value

        # 字符集，排序规则
        x_res[COLS_POS['字符集']] = process_special(r'CHARACTER-SET (\w+)', x)
        x_res[COLS_POS['排序规则']] = process_special(r'COLLATE (\w+)', x)

    print('处理后：🍼%s🍼' % x_res)
    print('')
    return x_res


def str_to_table(db, a_table, **kwargs):
    df = read(db_str=db, table_str='', sql_list='show CREATE TABLE %s' % a_table, **kwargs)
    table_name = df[0][0]
    rows = df[0][1].replace(',\n', '\n')
    rows = rows.split('\n')
    # tail = rows[-1]
    # print(tail)
    rows = rows[1:]
    rows = list(map(deal_a_row, rows))
    rows.insert(0, MY_COLS)
    rows.insert(0, ['Name', table_name])

    # 尝试寻找表的comment，当做是它的中文名（更加直观的名字）
    comment_of_this_table = ''
    pattern = 'COMMENT='
    if rows[-1][-1].startswith(pattern):  # 这里假设标准打印的语句的最后一行是包含了这个想要的comment
        comment_of_this_table = rows[-1][-1]
        print('got the comment of this table: %s' % comment_of_this_table)
        comment_of_this_table = comment_of_this_table.replace(pattern, '')
        del rows[-1][-1]
    print('finally, comment of this table: %s' % comment_of_this_table)
    if comment_of_this_table:
        rows.insert(1, ['COMMENT', comment_of_this_table])

    # DEBUG
    for a_row in rows:
        print(a_row)
    # assert False

    return {'data': rows, 'table_comment': comment_of_this_table}


def set_value_for_a_table(ws, t_data, row_num):
    max_row_len = 0  # 每一行最多有多少列
    data = t_data['data']
    len_data = len(data)
    print('总行数：%s' % len_data)

    table_comment = t_data['table_comment']
    print('table_comment: %s' % table_comment)
    print('表数据第一行为：%s' % data[0])

    for idx_row in range(len_data):
        row = data[idx_row]
        len_row = len(row)
        max_row_len = max(max_row_len, len_row)
        this_row_num = row_num + idx_row + 1
        for idx_col in range(len_row):
            this_cell = ws.cell(row=this_row_num, column=idx_col+1)
            assert ' ' not in row[idx_col]
            this_cell.value = row[idx_col]
            if idx_row == 0:  # 如果是第一行，这是表名
                this_cell.font = BOLD_RED_FONT
            else:
                this_cell.font = COMMON_CELL_FONT
        # if idx_row == 0:  # 如果是第一行，这是表名
        #     if len_row == 1:
        #         ws.merge_cells('A{0}:F{0}'.format(this_row_num))
        #     else:
        #         ws.merge_cells('A{0}:C{0}'.format(this_row_num))
        #         ws.merge_cells('D{0}:F{0}'.format(this_row_num))

    print('每一行最多有多少列：%s' % max_row_len)
    for idx_row in range(len_data):
        for idx_col in range(max_row_len):
            this_cell = ws.cell(row=row_num + idx_row + 1, column=idx_col + 1)

            # 样式设置
            # 位置设置
            this_cell.alignment = CENTER_H_V
            # 边界设置
            side = Side(border_style='thick', color="FF000000")
            border_upper_left = Border(left=side, top=side)
            border_upper_right = Border(right=side, top=side)
            border_bottom_left = Border(left=side, bottom=side)
            border_bottom_right = Border(right=side, bottom=side)
            border_left = Border(left=side)
            border_right = Border(right=side)
            border_top = Border(top=side)
            border_bottom = Border(bottom=side)
            if idx_row == 0:
                if idx_col == 0:
                    this_cell.border = border_upper_left
                elif idx_col == max_row_len - 1:
                    this_cell.border = border_upper_right
                else:
                    this_cell.border = border_top
            elif idx_row == len_data - 1:
                if idx_col == 0:
                    this_cell.border = border_bottom_left
                elif idx_col == max_row_len - 1:
                    this_cell.border = border_bottom_right
                else:
                    this_cell.border = border_bottom
            else:
                if idx_col == 0:
                    this_cell.border = border_left
                elif idx_col == max_row_len - 1:
                    this_cell.border = border_right
    return row_num+len_data


def main(file_name, db_s, **kwargs):
    print('working ...')

    if isinstance(db_s, str):
        db_s = [db_s]

    final_file_name = file_name
    if not os.path.isabs(file_name):
        print('输入的文件名不是绝对路径，尝试在前面加上base_dir')
        base_dir = kwargs.get('base_dir')
        if base_dir:
            final_file_name = os.path.join(base_dir, file_name)
            if not os.path.isabs(final_file_name):
                sys_exit('传入的base_dir加上也还不是绝对路径：%s' % base_dir)
        else:
            base_dir = BASE_DIR
            final_file_name = os.path.join(base_dir, file_name)
            if not os.path.isabs(final_file_name):
                sys_exit('配置中的BASE_DIR加上也还不是绝对路径：%s' % base_dir)
    print('获得最后的文件存储的绝对路径：%s' % final_file_name)

    for a_db in db_s:
        for_a_db(final_file_name, a_db, **kwargs)
    print('done ...')


def for_a_db(final_file_name, db, **kwargs):
    target_db_sheet_name = 'DB→%s' % db
    if os.path.exists(final_file_name):
        wb = openpyxl.load_workbook(final_file_name)
        # 如果 文件存在，那么寻找 目录sheet
        try:
            ws_catalog = wb['表目录']
            # TODO 删除原来的 表目录中的关于这个库的记录
        except Exception as err:
            print('不能够在以及存在的文件中找到 表目录')
            print('如果要继续，删除已经存在的文件，再次运行程序，同名文件将被自动创建')
            raise err
        try:
            # print(wb.sheetnames)
            del wb[target_db_sheet_name]  # 如果原来的文件中含有这个库的记录，删除了
        except KeyError as err:
            print('原文件中不存在 %s 的记录' % db)
            print(err)
        finally:
            ws_target_db_sheet = wb.create_sheet(target_db_sheet_name)
    else:
        wb = openpyxl.Workbook()
        del wb['Sheet']
        ws_catalog = wb.create_sheet('表目录')
        catalog_cols = list(CATALOG_COLS.keys())
        for idx in range(len(CATALOG_COLS)):
            this_cell = ws_catalog.cell(row=1, column=idx + 1)
            this_cell.value = catalog_cols[idx]
            this_cell.alignment = CENTER_H_V
        catalog_col_width = 33
        ws_catalog.column_dimensions['A'].width = catalog_col_width
        ws_catalog.column_dimensions['B'].width = catalog_col_width
        ws_catalog.column_dimensions['C'].width = catalog_col_width
        ws_target_db_sheet = wb.create_sheet(target_db_sheet_name)

    # 获取目前的catalog中的行数和列数
    print('表目录 页的维度：%s，最大行：%s，最大列：%s' % (ws_catalog.dimensions, ws_catalog.max_row, ws_catalog.max_column))

    all_tables = read(db_str=db, table_str='', sql_list='show tables', **kwargs)
    all_tables = list(map(lambda x: x[0], all_tables))
    len_all_tables = len(all_tables)
    print('在 %s 中，总共找到 %s 张表' % (db, len_all_tables))
    # assert False

    row_num = 0
    # for a_table in [['t_comb_mkt_estimation']]:  # DEBUG 单独拿一张表，进行测试
    # for a_table in all_tables.values[:11]:  # DEBUG 拿几张表，进行测试
    for a_table in all_tables:
        print('现在处理表：%s' % a_table)
        data = str_to_table(db, a_table, **kwargs)

        # 设置表目录中的对应信息
        # 把对应行数，第一列的单元格作为链接的目标单元格，添加到表目录中
        new_catalog_row_dict = {
            '库': db, '表': a_table, 'COMMENT': data['table_comment'],
            'row_num': ws_catalog.max_row + 1
        }
        for catalog_col in CATALOG_COLS:
            this_cell = ws_catalog.cell(row=new_catalog_row_dict['row_num'], column=CATALOG_COLS[catalog_col])
            this_cell.value = new_catalog_row_dict[catalog_col]
            this_cell.alignment = CENTER_H_V
            if catalog_col == '表':
                link = '#%s!A%s' % (target_db_sheet_name, row_num + 3)
                c_v = '=HYPERLINK(\"%s\", \"%s\")' % (link, new_catalog_row_dict[catalog_col])
                this_cell.value = c_v
                # this_cell.hyperlink = '#%s!A%s' % (target_db_sheet_name, row_num + 3)

        row_num = set_value_for_a_table(ws_target_db_sheet, data, row_num)
        row_num += 2  # 每一个表格之间 相隔两个空行

    # 把所有列的宽度设置为当列值长度最大的那一
    for column in ws_target_db_sheet.columns:
        col_count = ''
        max_len = 0
        for cell in column:
            # print(dir(cell))
            # assert False
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                pass
            else:
                col_count = cell.column_letter
                if cell.value:
                    max_len = max(len(cell.value), max_len)
        set_len = min(max_len * FONT_MM, 77)
        set_len = max(set_len, 8.88)
        print('列：%s的最大宽度：%s→%s' % (col_count, max_len, set_len))
        ws_target_db_sheet.column_dimensions[col_count].width = set_len
    # assert False

    wb.save(final_file_name)


def gui_main():
    # 类似java Frame的顶层容器
    base = tkinter.Tk()
    base.title('MySQL数据库导出')

    # 创建一个Label
    # 指定字体名称、大小、样式
    # ft = tk_font.Font(family='Fixdsys', size=10, weight=tk_font.BOLD)
    # ft_s10 = tk_font.Font(size=10)
    ft_s20 = tk_font.Font(size=20)

    lb_ip = tkinter.Label(base, text="服务器IP")
    lb_ip.pack()  # 布局设置，pack是一种自适应布局
    et_ip = tkinter.Entry(base, bd=5, font=ft_s20)
    et_ip.insert(0, '192.168.')
    et_ip.pack()

    lb_port = tkinter.Label(base, text="端口")
    lb_port.pack()
    et_port = tkinter.Entry(base, bd=5, font=ft_s20)
    et_port.insert(0, '3306')
    et_port.pack()

    lb_user = tkinter.Label(base, text="用户名")
    lb_user.pack()
    et_user = tkinter.Entry(base, bd=5, font=ft_s20)
    et_user.pack()

    lb_password = tkinter.Label(base, text="密码")
    lb_password.pack()
    et_password = tkinter.Entry(base, bd=5, font=ft_s20)
    et_password.pack()
    et_password["show"] = "*"

    lb_dbs = tkinter.Label(base, text=u"数据库列表\n注意：一个或多个，空格隔开")
    lb_dbs.pack()
    et_dbs = tkinter.Entry(base, bd=5, font=ft_s20)
    et_dbs.pack()

    # button测试
    def cmd1():
        info = {
            '服务器IP': et_ip.get(), '端口': et_port.get(), '用户名': et_user.get(), '密码': et_password.get(),
            '数据库列表': et_dbs.get()
        }
        for x in info:
            if info[x] == '':
                t_m.showerror(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.showinfo(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.showwarning(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.askretrycancel(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.askquestion(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.askyesno(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.askyesnocancel(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                # t_m.askokcancel(title='Attention', message='%s 的值是不是忘记输入啦' % x)
                return
        host = info['服务器IP']
        port = info['端口']
        try:
            port = int(port)
        except Exception as err:
            t_m.showerror(message='%s 不是int吧？↓%s' % (port, err))
            return

        user = info['用户名']
        password = info['密码']

        db_list = info['数据库列表']
        try:
            db_list = db_list.split()
            print('数据库列表:', db_list)
        except Exception as err:
            t_m.showerror(message=err)
            return

        file_name = '%s在服务器%s(%s)上的导出记录.xlsx' % (user, host, port)
        where_is_this_file = os.path.abspath(__file__)
        where_is_this_file = os.path.dirname(where_is_this_file)
        file_name = os.path.join(where_is_this_file, file_name)

        # 信息确认
        msg = [
            '请确认如下信息：',
            '服务器IP：%s' % host,
            '端口：%s' % port,
            '用户名：%s' % user,
            # '密码：%s' % password,
            '数据库列表：%s' % db_list,
        ]
        msg = '\n'.join(msg)
        answer = t_m.askyesno(message=msg)
        if not answer:
            return

        try:
            main(file_name=file_name, db_s=db_list, user=user, port=port, host=host, password=password)
        except Exception as err:
            t_m.showerror(message='导出失败↓\n%s' % err)
            return
        t_m.showinfo(message='导出完成，文件存储路径↓\n%s' % file_name)

    btn = tkinter.Button(base, text="导出", command=cmd1, font=ft_s20)
    btn.pack()

    # 消息循环，放在组件设置的最后
    base.mainloop()


if __name__ == '__main__':
    gui_main()
